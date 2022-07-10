import tensorflow as tf
import cv2
import os
import numpy as np
import time
import openpyxl
from resnet import resnet18, resnet10, resnet34, resnet48, resnet101

def main():

    ExcelPath = 'E:\\AOItest\\XLWchatu\\CapData3.xlsx'
    DataName = "Set14_1"
    Start_Col = 1 + 3*3    # 变换数据集
    Start_Row = 2 + 10*4   # 变换方法                                   # 0.  ！！！ 新的实验 这里是否要改
    workbook = openpyxl.load_workbook(ExcelPath)  # 新建
    worksheet = workbook.worksheets[6 - 1]
    ##### 加载权重文件                                                   # 1. ！！！新的实验  这里是否要改
    model_path = r'E:\AOItest\AOIpyOtherMethods\Cap_ResNet\\weights_resnet48_epo30_Data1\resnet18'

    # 创建新模型读取权重
    model = resnet18()                                                 #  2. ！！！新的实验   这里是否要改
    # 读取权重到新模型
    model.load_weights(model_path)
    # model.summary()  #输出 加载的网络信息
    # model.build(input_shape=[None, 64, 64, 3])

    # resNet18_40_Data1  0.005
    # resNet18_35_Data1  0.5  ***
    # resNet48_25_Data1  0.5
    # resNet48_28_Data1  0.8
    # resNet48_30_Data1  0.01(论文用)  0.5(better) ***
    # resNet48_50_Data1  0.5
    # resNet101_20_Data1 0.98 1.0    ***
    # resNet101_30_Data1 0.98 1.0    ***
    yuzhi = 0.01                                          #  3. ！！！注意 新的实验
    strat_time = time.time()
    # # res < 0.5 -->  up
    ImageFile1 = r"E:\\AOItest\\XiaoLunWenDataSet\\" + DataName + "\\up\\"
    ImageNames1 = os.listdir(ImageFile1)
    num = 0
    jishu = 0
    avg = 0.0
    for srcName in ImageNames1:
        src = get_img(ImageFile1+srcName)
        res = model.predict(src)
        res = round(float(res), 4)
        # print(res)
        avg = avg + res
        if res >= yuzhi:
            num = num + 1
        jishu = jishu + 1
        if jishu % 1000 == 0:
            print("已检测完", jishu)
    # print("--------------num", num)

    ImageFile2 = r"E:\\AOItest\\XiaoLunWenDataSet\\" + DataName + "\\Down\\"
    ImageNames2 = os.listdir(ImageFile2)
    num1 = 0
    avg2 = 0.0
    for srcName in ImageNames2:
        src = get_img(ImageFile2 + srcName)
        res = model.predict(src)
        res = round(float(res), 4)
        avg2 = avg2 + res
        # print(res)
        if res < yuzhi:
            num1 = num1 + 1

    avg = avg / len(ImageNames1)
    avg2 = avg2 / len(ImageNames2)
    print("avg ", avg, "avg2", avg2 )

    end_time = time.time()
    alltime = end_time - strat_time
    print(DataName + "已检测完成 耗时：" + str(end_time - strat_time))
    allnum = len(ImageNames1) + len(ImageNames2)
    avgtime = float(alltime / allnum)

    tp, fn, fp, tn = len(ImageNames1) - num, num, num1, len(ImageNames2) - num1
    accuracy = format((tp + tn) / (tp + fn + fp + tn), '.4f')  # 准确率
    # precision = format(tp / (tp + fp), '.4f')  # 精确率
    recall = format(tp / (tp + fn), '.4f')  # 召回率
    F1 = format((tp * 2) / (2 * tp + fn + fp), '.4f')  # F1值（H-mean值）
    MDR = format((fp) / (tp + fn + fp + tn), '.4f')  # F1值（H-mean值）


    worksheet.cell(Start_Row, Start_Col).value = "TP:"
    worksheet.cell(Start_Row+1, Start_Col).value = "FN:"
    worksheet.cell(Start_Row+2, Start_Col).value = "FP:"
    worksheet.cell(Start_Row+3, Start_Col).value = "TN:"

    worksheet.cell(Start_Row+4, Start_Col).value = "准确率:"
    worksheet.cell(Start_Row+5, Start_Col).value = "召回率:"
    worksheet.cell(Start_Row+6, Start_Col).value = "F1值:"
    worksheet.cell(Start_Row+7, Start_Col).value = "MDR:"
    worksheet.cell(Start_Row+8, Start_Col).value = "耗时:"

    worksheet.cell(Start_Row, Start_Col + 2).value = str(avgtime)
    worksheet.cell(Start_Row, Start_Col + 1).value = str(tp)
    worksheet.cell(Start_Row+1, Start_Col + 1).value = str(fn)
    worksheet.cell(Start_Row+2, Start_Col + 1).value = str(fp)
    worksheet.cell(Start_Row+3, Start_Col + 1).value = str(tn)
    worksheet.cell(Start_Row+4, Start_Col + 1).value = str(accuracy)
    worksheet.cell(Start_Row+5, Start_Col + 1).value = str(recall)
    worksheet.cell(Start_Row+6, Start_Col + 1).value = str(F1)
    worksheet.cell(Start_Row+7, Start_Col + 1).value = str(MDR)
    worksheet.cell(Start_Row+8, Start_Col + 1).value = str(end_time - strat_time)


    print(accuracy, '', recall, '', F1, '', MDR)
    workbook.save(ExcelPath)

    # prob = tf.nn.softmax(res, axis=1)
    # print(prob)
    # pred = tf.argmax(prob, axis=1)
    # pred = tf.cast(pred, dtype=tf.int32)
    # print(int(pred))

def get_img(data_path):
    # Getting image array from path:
    img = cv2.imread(data_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)     ## 这里 一定要要注意 提出的信息 要和 训练文件格式一样 RGB排列
    img = cv2.resize(img, (64, 64))
    img = img.reshape(1, 64, 64, 3)
    img = np.array(img / 255., dtype=np.float64)
    return img


if __name__ == "__main__":
    main()




